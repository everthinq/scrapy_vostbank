import json
import scrapy
from openpyxl import load_workbook


class vostbank(scrapy.Spider):
    name = 'vostbank'

    def start_requests(self):
        self.regions = [
            {"city_EN": "abakan", "city_id": "245", "city_RU": "Абакан"},
            {"city_EN": "aginskoe", "city_id": "131", "city_RU": "Агинское"},
            {"city_EN": "aldan", "city_id": "562797", "city_RU": "Алдан"},
            {"city_EN": "amursk", "city_id": "275", "city_RU": "Амурск"},
            {"city_EN": "anapa", "city_id": "45651", "city_RU": "Анапа"},
            {"city_EN": "angarsk", "city_id": "134", "city_RU": "Ангарск"},
            {"city_EN": "anzhero-sudzhensk", "city_id": "445763", "city_RU": "Анжеро-Судженск"},
            {"city_EN": "apatity", "city_id": "184", "city_RU": "Апатиты"},
            {"city_EN": "armavir", "city_id": "376459", "city_RU": "Армавир "},
            {"city_EN": "arsenev", "city_id": "213", "city_RU": "Арсеньев"},
            {"city_EN": "artem", "city_id": "212", "city_RU": "Артем"},
            {"city_EN": "arhangelsk", "city_id": "112", "city_RU": "Архангельск"},
            {"city_EN": "arhara", "city_id": "445537", "city_RU": "Архара"},
            {"city_EN": "askiz", "city_id": "444799", "city_RU": "Аскиз"},
            {"city_EN": "astrahan", "city_id": "376499", "city_RU": "Астрахань"},
            {"city_EN": "achinsk", "city_id": "165", "city_RU": "Ачинск"},
            {"city_EN": "bajkalsk", "city_id": "139", "city_RU": "Байкальск"},
            {"city_EN": "balashiha", "city_id": "95", "city_RU": "Балашиха"},
            {"city_EN": "baltijsk", "city_id": "445773", "city_RU": "Балтийск"},
            {"city_EN": "barnaul", "city_id": "99", "city_RU": "Барнаул"},
            {"city_EN": "belaya-kalitva", "city_id": "256", "city_RU": "Белая Калитва"},
            {"city_EN": "belgorod", "city_id": "33023", "city_RU": "Белгород"},
            {"city_EN": "belovo", "city_id": "241417", "city_RU": "Белово"},
            {"city_EN": "belogorsk", "city_id": "101", "city_RU": "Белогорск"},
            {"city_EN": "berdigestyah", "city_id": "444881", "city_RU": "Бердигестях"},
            {"city_EN": "berdsk", "city_id": "46537", "city_RU": "Бердск"},
            {"city_EN": "berezniki", "city_id": "201", "city_RU": "Березники"},
            {"city_EN": "bijsk", "city_id": "100", "city_RU": "Бийск"},
            {"city_EN": "bikin", "city_id": "279", "city_RU": "Бикин"},
            {"city_EN": "birobidzhan", "city_id": "280", "city_RU": "Биробиджан"},
            {"city_EN": "blagoveshhensk", "city_id": "563017", "city_RU": "Благовещенск"},
            {"city_EN": "bogotol", "city_id": "445675", "city_RU": "Боготол"},
            {"city_EN": "boguchany", "city_id": "445309", "city_RU": "Богучаны"},
            {"city_EN": "bodajbo", "city_id": "445179", "city_RU": "Бодайбо"},
            {"city_EN": "bolshoj-kamen", "city_id": "219", "city_RU": "Большой Камень"},
            {"city_EN": "borovichi", "city_id": "190", "city_RU": "Боровичи"},
            {"city_EN": "borogoncy", "city_id": "445453", "city_RU": "Борогонцы"},
            {"city_EN": "borodino", "city_id": "445160", "city_RU": "Бородино"},
            {"city_EN": "bohan", "city_id": "445567", "city_RU": "Бохан"},
            {"city_EN": "bratsk", "city_id": "132", "city_RU": "Братск"},
            {"city_EN": "bryansk", "city_id": "67", "city_RU": "Брянск"},
            {"city_EN": "buzuluk", "city_id": "445776", "city_RU": "Бузулук"},
            {"city_EN": "vanino", "city_id": "274", "city_RU": "Ванино"},
            {"city_EN": "velikie-luki", "city_id": "168357", "city_RU": "Великие Луки"},
            {"city_EN": "velikij-novgorod", "city_id": "188", "city_RU": "Великий Новгород"},
            {"city_EN": "vilyujsk", "city_id": "562798", "city_RU": "Вилюйск"},
            {"city_EN": "vilyuchinsk", "city_id": "155", "city_RU": "Вилючинск"},
            {"city_EN": "vihorevka", "city_id": "141", "city_RU": "Вихоревка"},
            {"city_EN": "vladivostok", "city_id": "76", "city_RU": "Владивосток"},
            {"city_EN": "vladimir", "city_id": "122", "city_RU": "Владимир"},
            {"city_EN": "volgograd", "city_id": "563015", "city_RU": "Волгоград"},
            {"city_EN": "vologda", "city_id": "126", "city_RU": "Вологда"},
            {"city_EN": "voronezh", "city_id": "68", "city_RU": "Воронеж"},
            {"city_EN": "vyborg", "city_id": "177", "city_RU": "Выборг"},
            {"city_EN": "vyazemskij", "city_id": "445873", "city_RU": "Вяземский"},
            {"city_EN": "gatchina", "city_id": "175", "city_RU": "Гатчина"},
            {"city_EN": "gubkin", "city_id": "33017", "city_RU": "Губкин"},
            {"city_EN": "gusev", "city_id": "445285", "city_RU": "Гусев"},
            {"city_EN": "dalnegorsk", "city_id": "220", "city_RU": "Дальнегорск"},
            {"city_EN": "dalnerechensk", "city_id": "214", "city_RU": "Дальнереченск"},
            {"city_EN": "dzerzhinsk", "city_id": "168351", "city_RU": "Дзержинск"},
            {"city_EN": "divnogorsk", "city_id": "171", "city_RU": "Дивногорск"},
            {"city_EN": "dolinsk", "city_id": "266", "city_RU": "Долинск"},
            {"city_EN": "egorevsk", "city_id": "445788", "city_RU": "Егорьевск"},
            {"city_EN": "ekaterinburg", "city_id": "268", "city_RU": "Екатеринбург"},
            {"city_EN": "elizovo", "city_id": "154", "city_RU": "Елизово"},
            {"city_EN": "enisejsk", "city_id": "444812", "city_RU": "Енисейск"},
            {"city_EN": "zheleznogorsk", "city_id": "162", "city_RU": "Железногорск"},
            {"city_EN": "zheleznogorsk-ilimskij", "city_id": "137", "city_RU": "Железногорск-Илимский"},
            {"city_EN": "zheleznodorozhnyj", "city_id": "376477", "city_RU": "Железнодорожный"},
            {"city_EN": "zavitinsk", "city_id": "111", "city_RU": "Завитинск"},
            {"city_EN": "zaigraevo", "city_id": "445120", "city_RU": "Заиграево"},
            {"city_EN": "zalari", "city_id": "445862", "city_RU": "Залари"},
            {"city_EN": "zaozernyj", "city_id": "445410", "city_RU": "Заозерный"},
            {"city_EN": "zelenogorsk", "city_id": "164", "city_RU": "Зеленогорск"},
            {"city_EN": "zelenograd", "city_id": "376481", "city_RU": "Зеленоград"},
            {"city_EN": "zeya", "city_id": "104", "city_RU": "Зея"},
            {"city_EN": "zima", "city_id": "142", "city_RU": "Зима"},
            {"city_EN": "zlatoust", "city_id": "376463", "city_RU": "Златоуст"},
            {"city_EN": "ivanovo", "city_id": "376489", "city_RU": "Иваново"},
            {"city_EN": "izhevsk", "city_id": "588715", "city_RU": "Ижевск"},
            {"city_EN": "ilanskij", "city_id": "445800", "city_RU": "Иланский"},
            {"city_EN": "irkutsk", "city_id": "70", "city_RU": "Иркутск"},
            {"city_EN": "iskitim", "city_id": "33011", "city_RU": "Искитим"},
            {"city_EN": "ishim", "city_id": "445777", "city_RU": "Ишим"},
            {"city_EN": "kavalerovo", "city_id": "445865", "city_RU": "Кавалерово"},
            {"city_EN": "kazan", "city_id": "168359", "city_RU": "Казань"},
            {"city_EN": "kaliningrad", "city_id": "150", "city_RU": "Калининград"},
            {"city_EN": "kaluga", "city_id": "71", "city_RU": "Калуга"},
            {"city_EN": "kansk", "city_id": "170", "city_RU": "Канск"},
            {"city_EN": "kachug", "city_id": "445063", "city_RU": "Качуг"},
            {"city_EN": "kemerovo", "city_id": "157", "city_RU": "Кемерово"},
            {"city_EN": "kingisepp", "city_id": "178", "city_RU": "Кингисепп"},
            {"city_EN": "kirensk", "city_id": "445167", "city_RU": "Киренск"},
            {"city_EN": "kirov", "city_id": "376473", "city_RU": "Киров"},
            {"city_EN": "kiselevsk", "city_id": "376281", "city_RU": "Киселевск"},
            {"city_EN": "kislovodsk", "city_id": "376469", "city_RU": "Кисловодск"},
            {"city_EN": "klincy", "city_id": "168353", "city_RU": "Клинцы"},
            {"city_EN": "kovrov", "city_id": "121", "city_RU": "Ковров"},
            {"city_EN": "kolomna", "city_id": "376483", "city_RU": "Коломна "},
            {"city_EN": "kolpino", "city_id": "174", "city_RU": "Колпино"},
            {"city_EN": "komsomolsk-na-amure", "city_id": "84", "city_RU": "Комсомольск-на-Амуре"},
            {"city_EN": "kondopoga", "city_id": "445742", "city_RU": "Кондопога"},
            {"city_EN": "konosha", "city_id": "445444", "city_RU": "Коноша"},
            {"city_EN": "konstantinovka", "city_id": "445251", "city_RU": "Константиновка"},
            {"city_EN": "korsakov", "city_id": "261", "city_RU": "Корсаков"},
            {"city_EN": "kostroma", "city_id": "376497", "city_RU": "Кострома"},
            {"city_EN": "kotlas", "city_id": "115", "city_RU": "Котлас"},
            {"city_EN": "krasnogorsk", "city_id": "376479", "city_RU": "Красногорск"},
            {"city_EN": "krasnodar", "city_id": "632890", "city_RU": "Краснодар"},
            {"city_EN": "krasnokamensk", "city_id": "128", "city_RU": "Краснокаменск"},
            {"city_EN": "krasnokamsk", "city_id": "205", "city_RU": "Краснокамск"},
            {"city_EN": "krasnoyarsk", "city_id": "73", "city_RU": "Красноярск"},
            {"city_EN": "kudymkar", "city_id": "203", "city_RU": "Кудымкар"},
            {"city_EN": "kujtun", "city_id": "445875", "city_RU": "Куйтун"},
            {"city_EN": "kungur", "city_id": "204", "city_RU": "Кунгур"},
            {"city_EN": "kurgan", "city_id": "168345", "city_RU": "Курган"},
            {"city_EN": "kurovskoe", "city_id": "96", "city_RU": "Куровское"},
            {"city_EN": "kursk", "city_id": "173", "city_RU": "Курск"},
            {"city_EN": "kyzyl", "city_id": "244", "city_RU": "Кызыл"},
            {"city_EN": "kyahta", "city_id": "444871", "city_RU": "Кяхта"},
            {"city_EN": "leninsk-kuzneckij", "city_id": "445095", "city_RU": "Ленинск-Кузнецкий"},
            {"city_EN": "leninskoe", "city_id": "445042", "city_RU": "Ленинское"},
            {"city_EN": "lensk", "city_id": "562799", "city_RU": "Ленск"},
            {"city_EN": "lesozavodsk", "city_id": "218", "city_RU": "Лесозаводск"},
            {"city_EN": "lesosibirsk", "city_id": "166", "city_RU": "Лесосибирск"},
            {"city_EN": "lipeck", "city_id": "181", "city_RU": "Липецк"},
            {"city_EN": "luga", "city_id": "180", "city_RU": "Луга"},
            {"city_EN": "luchegorsk", "city_id": "215", "city_RU": "Лучегорск"},
            {"city_EN": "lgov", "city_id": "445145", "city_RU": "Льгов"},
            {"city_EN": "lyubercy", "city_id": "44231", "city_RU": "Люберцы"},
            {"city_EN": "magadan", "city_id": "33009", "city_RU": "Магадан"},
            {"city_EN": "magdagachi", "city_id": "444995", "city_RU": "Магдагачи"},
            {"city_EN": "magistralnyj", "city_id": "444863", "city_RU": "Магистральный"},
            {"city_EN": "magnitogorsk", "city_id": "241423", "city_RU": "Магнитогорск"},
            {"city_EN": "majya", "city_id": "445459", "city_RU": "Майя"},
            {"city_EN": "mariinsk", "city_id": "159", "city_RU": "Мариинск"},
            {"city_EN": "medvezhegorsk", "city_id": "444798", "city_RU": "Медвежьегорск"},
            {"city_EN": "mezhdurechensk", "city_id": "562796", "city_RU": "Междуреченск"},
            {"city_EN": "minusinsk", "city_id": "163", "city_RU": "Минусинск"},
            {"city_EN": "mirnyj", "city_id": "562800", "city_RU": "Мирный"},
            {"city_EN": "monchegorsk", "city_id": "182", "city_RU": "Мончегорск"},
            {"city_EN": "moskva", "city_id": "54", "city_RU": "Москва"},
            {"city_EN": "murmansk", "city_id": "74", "city_RU": "Мурманск"},
            {"city_EN": "murom", "city_id": "445782", "city_RU": "Муром"},
            {"city_EN": "naberezhnye-chelny", "city_id": "168361", "city_RU": "Набережные Челны"},
            {"city_EN": "nazarovo", "city_id": "167", "city_RU": "Назарово"},
            {"city_EN": "namcy", "city_id": "445002", "city_RU": "Намцы"},
            {"city_EN": "nahodka", "city_id": "78", "city_RU": "Находка"},
            {"city_EN": "nevelsk", "city_id": "267", "city_RU": "Невельск"},
            {"city_EN": "nevinnomyssk", "city_id": "445831", "city_RU": "Невинномысск"},
            {"city_EN": "nerchinsk", "city_id": "444950", "city_RU": "Нерчинск"},
            {"city_EN": "neryungri", "city_id": "610526", "city_RU": "Нерюнгри"},
            {"city_EN": "nizhnevartovsk", "city_id": "33019", "city_RU": "Нижневартовск"},
            {"city_EN": "nizhnekamsk", "city_id": "445204", "city_RU": "Нижнекамск"},
            {"city_EN": "nizhneudinsk", "city_id": "143", "city_RU": "Нижнеудинск"},
            {"city_EN": "nizhnij-bestyah", "city_id": "445523", "city_RU": "Нижний Бестях"},
            {"city_EN": "nizhnij-novgorod", "city_id": "185", "city_RU": "Нижний Новгород"},
            {"city_EN": "nizhnij-tagil", "city_id": "376453", "city_RU": "Нижний Тагил"},
            {"city_EN": "nikolaevsk-na-amure", "city_id": "277", "city_RU": "Николаевск-на-Амуре"},
            {"city_EN": "novokuzneck", "city_id": "156", "city_RU": "Новокузнецк"},
            {"city_EN": "novorossijsk", "city_id": "376457", "city_RU": "Новороссийск "},
            {"city_EN": "novosibirsk", "city_id": "98", "city_RU": "Новосибирск"},
            {"city_EN": "novocherkassk", "city_id": "376475", "city_RU": "Новочеркасск"},
            {"city_EN": "nogliki", "city_id": "265", "city_RU": "Ноглики"},
            {"city_EN": "norilsk", "city_id": "161", "city_RU": "Норильск"},
            {"city_EN": "nyurba", "city_id": "445439", "city_RU": "Нюрба"},
            {"city_EN": "obluche", "city_id": "282", "city_RU": "Облучье"},
            {"city_EN": "obninsk", "city_id": "445787", "city_RU": "Обнинск"},
            {"city_EN": "odincovo", "city_id": "376485", "city_RU": "Одинцово"},
            {"city_EN": "omsk", "city_id": "191", "city_RU": "Омск"},
            {"city_EN": "orel", "city_id": "563018", "city_RU": "Орел"},
            {"city_EN": "orenburg", "city_id": "45675", "city_RU": "Оренбург"},
            {"city_EN": "orsk", "city_id": "241433", "city_RU": "Орск"},
            {"city_EN": "osa", "city_id": "445729", "city_RU": "Оса"},
            {"city_EN": "oha", "city_id": "263", "city_RU": "Оха"},
            {"city_EN": "partizansk", "city_id": "217", "city_RU": "Партизанск"},
            {"city_EN": "penza", "city_id": "445732", "city_RU": "Пенза"},
            {"city_EN": "pervomajskij", "city_id": "445600", "city_RU": "Первомайский"},
            {"city_EN": "pereyaslavka", "city_id": "445867", "city_RU": "Переяславка"},
            {"city_EN": "perm", "city_id": "75", "city_RU": "Пермь"},
            {"city_EN": "pestovo", "city_id": "445599", "city_RU": "Пестово"},
            {"city_EN": "petrozavodsk", "city_id": "233", "city_RU": "Петрозаводск"},
            {"city_EN": "petropavlovka", "city_id": "444894", "city_RU": "Петропавловка"},
            {"city_EN": "petropavlovsk-kamchatskij", "city_id": "72", "city_RU": "Петропавловск-Камчатский"},
            {"city_EN": "pleseck", "city_id": "445508", "city_RU": "Плесецк"},
            {"city_EN": "podolsk", "city_id": "89", "city_RU": "Подольск"},
            {"city_EN": "pokrovka", "city_id": "445409", "city_RU": "Покровка"},
            {"city_EN": "pokrovsk", "city_id": "445872", "city_RU": "Покровск"},
            {"city_EN": "poronajsk", "city_id": "264", "city_RU": "Поронайск"},
            {"city_EN": "pskov", "city_id": "79", "city_RU": "Псков"},
            {"city_EN": "pyatigorsk", "city_id": "376467", "city_RU": "Пятигорск"},
            {"city_EN": "rajchihinsk", "city_id": "108", "city_RU": "Райчихинск"},
            {"city_EN": "rostov-na-donu", "city_id": "81", "city_RU": "Ростов-на-Дону"},
            {"city_EN": "ryazan", "city_id": "376495", "city_RU": "Рязань"},
            {"city_EN": "samara", "city_id": "45655", "city_RU": "Самара"},
            {"city_EN": "sankt-peterburg", "city_id": "55", "city_RU": "Санкт-Петербург"},
            {"city_EN": "saransk", "city_id": "240", "city_RU": "Саранск"},
            {"city_EN": "saratov", "city_id": "44223", "city_RU": "Саратов"},
            {"city_EN": "sayanogorsk", "city_id": "246", "city_RU": "Саяногорск"},
            {"city_EN": "sayansk", "city_id": "135", "city_RU": "Саянск"},
            {"city_EN": "svetlyj", "city_id": "253", "city_RU": "Светлый"},
            {"city_EN": "svobodnyj", "city_id": "103", "city_RU": "Свободный"},
            {"city_EN": "severobajkalsk", "city_id": "230", "city_RU": "Северобайкальск"},
            {"city_EN": "severodvinsk", "city_id": "113", "city_RU": "Северодвинск"},
            {"city_EN": "severomorsk", "city_id": "183", "city_RU": "Североморск"},
            {"city_EN": "selenginsk", "city_id": "232", "city_RU": "Селенгинск"},
            {"city_EN": "sergiev-posad", "city_id": "91", "city_RU": "Сергиев Посад"},
            {"city_EN": "serov", "city_id": "376455", "city_RU": "Серов"},
            {"city_EN": "serpuhov", "city_id": "86", "city_RU": "Серпухов"},
            {"city_EN": "seryshevo", "city_id": "107", "city_RU": "Серышево"},
            {"city_EN": "skovorodino", "city_id": "106", "city_RU": "Сковородино"},
            {"city_EN": "slavyanka", "city_id": "445845", "city_RU": "Славянка"},
            {"city_EN": "slavyansk-na-kubani", "city_id": "376461", "city_RU": "Славянск-на-Кубани"},
            {"city_EN": "slancy", "city_id": "176", "city_RU": "Сланцы"},
            {"city_EN": "slyudyanka", "city_id": "140", "city_RU": "Слюдянка"},
            {"city_EN": "smolensk", "city_id": "269", "city_RU": "Смоленск"},
            {"city_EN": "sovetsk", "city_id": "33015", "city_RU": "Советск"},
            {"city_EN": "sovetskaya-gavan", "city_id": "276", "city_RU": "Советская Гавань"},
            {"city_EN": "solikamsk", "city_id": "206", "city_RU": "Соликамск"},
            {"city_EN": "solnechnyj", "city_id": "445289", "city_RU": "Солнечный"},
            {"city_EN": "sosnovoborsk", "city_id": "172", "city_RU": "Сосновоборск"},
            {"city_EN": "sosnovyj-bor", "city_id": "85", "city_RU": "Сосновый Бор"},
            {"city_EN": "sochi", "city_id": "243127", "city_RU": "Сочи"},
            {"city_EN": "spassk-dalnij", "city_id": "216", "city_RU": "Спасск-Дальний"},
            {"city_EN": "stavropol", "city_id": "45721", "city_RU": "Ставрополь"},
            {"city_EN": "staraya-russa", "city_id": "189", "city_RU": "Старая Русса"},
            {"city_EN": "staryj-oskol", "city_id": "117", "city_RU": "Старый Оскол"},
            {"city_EN": "sterlitamak", "city_id": "376449", "city_RU": "Стерлитамак"},
            {"city_EN": "surgut", "city_id": "620381", "city_RU": "Сургут"},
            {"city_EN": "syktyvkar", "city_id": "235", "city_RU": "Сыктывкар"},
            {"city_EN": "taganrog", "city_id": "241421", "city_RU": "Таганрог "},
            {"city_EN": "tajshet", "city_id": "144", "city_RU": "Тайшет"},
            {"city_EN": "tambov", "city_id": "241419", "city_RU": "Тамбов"},
            {"city_EN": "tambovka", "city_id": "445253", "city_RU": "Тамбовка"},
            {"city_EN": "tver", "city_id": "241425", "city_RU": "Тверь"},
            {"city_EN": "tihvin", "city_id": "168347", "city_RU": "Тихвин"},
            {"city_EN": "tobolsk", "city_id": "445813", "city_RU": "Тобольск"},
            {"city_EN": "tomsk", "city_id": "272", "city_RU": "Томск"},
            {"city_EN": "tosno", "city_id": "445688", "city_RU": "Тосно"},
            {"city_EN": "trudovoe", "city_id": "445859", "city_RU": "Трудовое"},
            {"city_EN": "tula", "city_id": "376491", "city_RU": "Тула"},
            {"city_EN": "tulun", "city_id": "146", "city_RU": "Тулун"},
            {"city_EN": "tynda", "city_id": "102", "city_RU": "Тында"},
            {"city_EN": "tyumen", "city_id": "23313", "city_RU": "Тюмень"},
            {"city_EN": "uglegorsk", "city_id": "105", "city_RU": "Углегорск"},
            {"city_EN": "uzhur", "city_id": "445101", "city_RU": "Ужур"},
            {"city_EN": "ulan-ude", "city_id": "229", "city_RU": "Улан-Удэ"},
            {"city_EN": "ulyanovsk", "city_id": "168365", "city_RU": "Ульяновск"},
            {"city_EN": "usole-sibirskoe", "city_id": "136", "city_RU": "Усолье-Сибирское"},
            {"city_EN": "ussurijsk", "city_id": "77", "city_RU": "Уссурийск"},
            {"city_EN": "ust-ilimsk", "city_id": "133", "city_RU": "Усть-Илимск"},
            {"city_EN": "ust-kut", "city_id": "147", "city_RU": "Усть-Кут"},
            {"city_EN": "ust-ordynskij", "city_id": "445047", "city_RU": "Усть-Ордынский"},
            {"city_EN": "ufa", "city_id": "445807", "city_RU": "Уфа"},
            {"city_EN": "uhta", "city_id": "236", "city_RU": "Ухта"},
            {"city_EN": "fokino", "city_id": "221", "city_RU": "Фокино"},
            {"city_EN": "xabarovsk", "city_id": "83", "city_RU": "Хабаровск"},
            {"city_EN": "ximki", "city_id": "376487", "city_RU": "Химки"},
            {"city_EN": "xolmsk", "city_id": "262", "city_RU": "Холмск"},
            {"city_EN": "xorol", "city_id": "445837", "city_RU": "Хороль"},
            {"city_EN": "chajkovskij", "city_id": "202", "city_RU": "Чайковский"},
            {"city_EN": "cheboksary", "city_id": "235311", "city_RU": "Чебоксары"},
            {"city_EN": "chegdomyn", "city_id": "278", "city_RU": "Чегдомын"},
            {"city_EN": "chelyabinsk", "city_id": "45671", "city_RU": "Челябинск"},
            {"city_EN": "cheremhovo", "city_id": "145", "city_RU": "Черемхово"},
            {"city_EN": "cherepovec", "city_id": "125", "city_RU": "Череповец"},
            {"city_EN": "chernigovka", "city_id": "445351", "city_RU": "Черниговка"},
            {"city_EN": "chernogorsk", "city_id": "247", "city_RU": "Черногорск"},
            {"city_EN": "chernushka", "city_id": "445760", "city_RU": "Чернушка"},
            {"city_EN": "chernyahovsk", "city_id": "152", "city_RU": "Черняховск"},
            {"city_EN": "chita", "city_id": "69", "city_RU": "Чита"},
            {"city_EN": "chuguevka", "city_id": "444782", "city_RU": "Чугуевка"},
            {"city_EN": "chunskij", "city_id": "597152", "city_RU": "Чунский"},
            {"city_EN": "churapcha", "city_id": "445380", "city_RU": "Чурапча"},
            {"city_EN": "sharypovo", "city_id": "444789", "city_RU": "Шарыпово"},
            {"city_EN": "shebekino", "city_id": "33025", "city_RU": "Шебекино"},
            {"city_EN": "shelehov", "city_id": "138", "city_RU": "Шелехов"},
            {"city_EN": "shimanovsk", "city_id": "109", "city_RU": "Шимановск"},
            {"city_EN": "shhelkovo", "city_id": "168349", "city_RU": "Щелково"},
            {"city_EN": "ytyk-kyuel", "city_id": "444960", "city_RU": "Ытык-Кюель"},
            {"city_EN": "elektrostal", "city_id": "90", "city_RU": "Электросталь"},
            {"city_EN": "engels", "city_id": "376451", "city_RU": "Энгельс"},
            {"city_EN": "yuzhno-sahalinsk", "city_id": "82", "city_RU": "Южно-Сахалинск"},
            {"city_EN": "yurga", "city_id": "33013", "city_RU": "Юрга"},
            {"city_EN": "yakutsk", "city_id": "556123", "city_RU": "Якутск"},
            {"city_EN": "yaroslavl", "city_id": "243113", "city_RU": "Ярославль"},
            {"city_EN": "yasnogorsk", "city_id": "130", "city_RU": "Ясногорск"}
        ]

        self.cell_value = '2'
        self.workbook = load_workbook('43. Public Stock Company Orient Express Bank Russia.xlsx')
        self.worksheet = self.workbook[self.workbook.sheetnames[0]]

        for region in self.regions:
            yield scrapy.Request(
                url='https://www.vostbank.ru/office/' + region['city_EN'] + '/?city_id=' + region['city_id'],
                callback=self.before_parse_json,
                meta={'city_RU': region['city_RU']}
            )
            # break

    def before_parse_json(self, response):
        cookies = response.headers.getlist('Set-Cookie')
        # print(cookies)

        city_id = str(cookies).split(',')[1].split(';')[0].strip()
        if 'city_id' not in city_id:
            city_id = str(cookies).split(',')[0].split(';')[0].strip()
        city_id = city_id.split('=')[1]

        city_name = str(cookies).split(',')[3].split(';')[0].strip()
        if 'city_name' not in city_name:
            city_name = str(cookies).split(',')[2].split(';')[0].strip()
        city_name = city_name.split('=')[1]

        city_kladr_id = str(cookies).split(',')[7].split(';')[0].strip()
        if 'city_kladr_id' not in city_kladr_id:
            city_kladr_id = str(cookies).split(',')[6].split(';')[0].strip()
        city_kladr_id = city_kladr_id.split('=')[1]

        cookies = {
            'city_kladr_id': city_kladr_id,
            'city_id': city_id,
            'city_name': city_name
        }

        yield scrapy.Request(
            url='https://www.vostbank.ru/local/templates/vostbank2019/ajax/api.front.php?type=getAtmsAndOffices&filter_type=only-office',
            cookies=cookies,
            callback=self.parse_json,
            dont_filter=True,
            meta={'city_RU': response.meta['city_RU']}
        )

    def parse_json(self, response):
        if (response.text):
            JSON = json.loads(response.text)
            # print(JSON)

            for branch in JSON['markers']:
                branch_name = branch['content']['title']
                address = branch['content']['address']
                city = response.meta['city_RU']
                lat = branch['lat']
                lng = branch['lng']

                print('Writing --', 'name:', branch_name, 'address:', address, 'city:', city, 'lat:', lat, 'lng:', lng)

                self.cell_value = str(self.cell_value)
                self.worksheet['B' + self.cell_value] = 'Public Stock Company Orient Express Bank Russia'
                self.worksheet['C' + self.cell_value] = branch_name
                self.worksheet['D' + self.cell_value] = address
                self.worksheet['G' + self.cell_value] = 'Russia'
                self.worksheet['H' + self.cell_value] = 'RU'
                self.worksheet['J' + self.cell_value] = city
                self.worksheet['M' + self.cell_value] = lat
                self.worksheet['N' + self.cell_value] = lng
                self.worksheet['O' + self.cell_value] = 'Address'
                self.worksheet['R' + self.cell_value] = 'Bank website'
                self.cell_value = int(self.cell_value) + 1

        self.workbook.save('43. Public Stock Company Orient Express Bank Russia.xlsx')
